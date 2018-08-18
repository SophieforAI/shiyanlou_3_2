# -*- coding: utf-8 -*-

from openpyxl import load_workbook # 可以用来载入已有数据表格
from openpyxl import Workbook # 可以用来处理新的数据表格
from datetime import datetime # 可以用来处理时间相关的数据
import pandas as pd
import os

def combine():
    '''
    1 读取数据
    2 合并数据
    3 写入combine表中
    4 保存原数据文件
    '''
    wb = load_workbook('E:/programs/shiyanlou/courses.xlsx')
    ws = wb.create_sheet(title="combine",index=3)

    student = list(wb['students'].values)
    row_name_s = student[0]
    student_frame = pd.DataFrame(student[1:],columns=row_name_s)

    time = list(wb['time'].values)
    row_name_t = time[0]
    time_frame = pd.DataFrame(time[1:],columns=row_name_t)

    combine_frame = pd.merge(student_frame,time_frame,on="课程名称",how='left')
    combine_frame = combine_frame[['创建时间_x','课程名称','学习人数','学习时间']]
    combine_name = ['创建时间','课程名称','学习人数','学习时间']
    combine_final_frame = [combine_name]+combine_frame.values.tolist()
    combine_sheet = wb['combine']
    for row in combine_final_frame:
        combine_sheet.append(row)

    wb.save('E:/programs/shiyanlou/combine.xlsx')




def split():
    '''
    该函数可以用来分割文件：
    1. 读取 combine 表中的数据
    2. 将数据按时间分割
    3. 写入不同的数据表中
    '''
    # 1 读取combine表
    combine = load_workbook('E:/programs/shiyanlou/combine.xlsx')
    combine_sheet = list(combine['combine'].values)
    combine_name = combine_sheet[0]
    combine_frame = pd.DataFrame(combine_sheet[1:],columns=combine_name)
    combine_frame['year'] = list(map(lambda x:x.year, combine_frame['创建时间']))
    year_set = list(set(combine_frame['year']))
    for year in year_set:
        combine_year =  combine_frame[combine_frame['year']==year]
        combine_year_sheet = combine_year.drop(columns='year')
        combine_final_frame = [combine_name] + combine_year_sheet.values.tolist()
        wb = Workbook(write_only=True)
        ws = wb.create_sheet()
        for row in combine_final_frame:
            ws.append(row)
        path = 'E:/programs/shiyanlou/{}.xlsx'.format(year)
        wb.save(path)







# 执行
if __name__ == '__main__':
    combine()
    split()